#!/usr/bin/env python3
import openpyxl
import json
import sys

try:
    # Load the workbook
    filepath = 'controll offers.xlsx'
    print(f'📂 Reading file: {filepath}\n')

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    print(f'📄 Sheets found: {", ".join(wb.sheetnames)}\n')

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f'\n📋 Sheet: {sheet_name}')
        print('─' * 80)

        # Get all rows
        rows = list(sheet.iter_rows(values_only=True))

        if len(rows) == 0:
            print('  Empty sheet')
            continue

        # First row is headers
        headers = rows[0]
        print(f'  Columns: {", ".join([str(h) for h in headers if h is not None])}')
        print(f'  Total rows (including header): {len(rows)}')

        # Show first 10 data rows
        print(f'\n  First 10 rows:')
        for i, row in enumerate(rows[1:11], 1):
            print(f'\n  {i}. {dict(zip(headers, row))}')

        # Extract all data for analysis
        data = []
        for row in rows[1:]:
            row_dict = dict(zip(headers, row))
            data.append(row_dict)

        # Save to JSON for easier processing
        output_file = f'{sheet_name.replace(" ", "_")}_data.json'
        with open(output_file, 'w') as f:
            json.dump(data, f, indent=2, default=str)
        print(f'\n  ✅ Data exported to: {output_file}')

    print('\n✅ Done!')

except Exception as e:
    print(f'❌ Error: {e}')
    import traceback
    traceback.print_exc()
    sys.exit(1)
